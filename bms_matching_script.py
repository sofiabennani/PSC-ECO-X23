
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

df_multi_bms = pd.read_excel("Final_Matching_with_BMS_colored.xlsx")

# On identifie le BMS le plus fréquent
bms_group = (
    df_multi_bms.groupby(['cnit', 'plant_bms_country', 'plant_bms_city'])
    .size()
    .reset_index(name='count')
)
most_common_bms = (
    bms_group.sort_values('count', ascending=False)
    .drop_duplicates(subset=['cnit'])
    .drop(columns='count')
)

# On ne garde qu'un seul BMS par CNIT
df_cleaned = df_multi_bms.drop(columns=['plant_bms_country', 'plant_bms_city'])
df_unique_bms = df_cleaned.merge(most_common_bms, on='cnit', how='left')
df_unique_bms = df_unique_bms.drop_duplicates()

#Exporter vers Excel
export_path = "Final_Matching_with_BMS_unique_per_cnit_colored_deduplicated.xlsx"
df_unique_bms.to_excel(export_path, index=False)
wb = load_workbook(export_path)
ws = wb.active

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

header = [cell.value for cell in ws[1]]
highlight_idx = header.index("highlight") + 1

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    value = row[highlight_idx - 1].value
    if value == "green":
        for cell in row:
            cell.fill = green_fill
    elif value == "red":
        for cell in row:
            cell.fill = red_fill

wb.save("Final_Matching_with_BMS_unique_per_cnit_colored_deduplicated_colored.xlsx")
